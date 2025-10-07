"""AWS Account Cost Analyzer

Produces a simple cost report using Cost Explorer and resource inventory APIs.

Run with --dry-run to avoid AWS calls and produce sample data.
"""
import argparse
import datetime
import json
import os
import sys
from collections import defaultdict

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError
except Exception:
    boto3 = None

# Global boto3 session (set from CLI --profile if provided)
_BOTO_SESSION = None

def get_session():
    global _BOTO_SESSION
    if _BOTO_SESSION:
        return _BOTO_SESSION
    if boto3:
        return boto3.Session()
    return None

from jinja2 import Template
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
except Exception:
    # ReportLab optional
    letter = None


SAMPLE_DATA = {
    "costs": [
        {"date": "2025-04-01", "amount": 120.0, "service": "AmazonEC2"},
        {"date": "2025-05-01", "amount": 150.0, "service": "AmazonEC2"},
        {"date": "2025-06-01", "amount": 200.0, "service": "AmazonEC2"},
        {"date": "2025-04-01", "amount": 80.0, "service": "AmazonS3"},
        {"date": "2025-05-01", "amount": 90.0, "service": "AmazonS3"},
        {"date": "2025-06-01", "amount": 110.0, "service": "AmazonS3"},
        {"date": "2025-04-01", "amount": 40.0, "service": "AmazonRDS"},
        {"date": "2025-05-01", "amount": 45.0, "service": "AmazonRDS"},
        {"date": "2025-06-01", "amount": 60.0, "service": "AmazonRDS"},
    ],
    "resources": {
        "EC2": 3,
        "RDS": 1,
        "S3_buckets": 5,
    },
    "resource_lists": {
        "EC2_list": ["i-0123456789abcdef0", "i-0fedcba9876543210", "i-00aa11bb22cc33dd4"],
        "EBS_volumes_list": ["vol-0123", "vol-0456"],
        "RDS_list": ["db-1"],
        "S3_buckets_list": ["my-bucket-a", "my-bucket-b"],
        "Lambda_functions_list": ["my-func-a"],
        "ECR_repos_list": ["repo-a"],
        "EKS_clusters_list": ["cluster-a"],
        "SQS_queues_list": ["https://sqs.us-east-1.amazonaws.com/123456789012/my-queue"]
    }
}


def parse_args():
    p = argparse.ArgumentParser(description="AWS Account Cost Analyzer")
    p.add_argument("--output", "-o", default="report.html", help="HTML report path")
    p.add_argument("--csv", default=None, help="Write CSV of cost time series")
    p.add_argument("--months", type=int, default=6, help="Months of history to analyze")
    p.add_argument("--forecast_months", type=int, default=3, help="Months to forecast forward")
    p.add_argument("--dry-run", action="store_true", help="Do not call AWS; use sample data")
    p.add_argument("--pdf", default=None, help="Write PDF report to this path")
    p.add_argument("--all-regions", action="store_true", help="Check all AWS regions for regional resources")
    p.add_argument("--regions", default=None, help="Comma-separated list of regions to check (overrides --all-regions)")
    p.add_argument("--concurrency", type=int, default=6, help="Max concurrency for region checks")
    p.add_argument("--tag-keys", default=None, help="Comma-separated tag keys to aggregate costs by (optional)")
    p.add_argument("--cur-file", default=None, help="Path to local CUR CSV (gz allowed) to parse for per-resource costs")
    p.add_argument("--cur-s3-bucket", default=None, help="S3 bucket name where CUR is stored (optional)")
    p.add_argument("--cur-prefix", default=None, help="S3 prefix/path to CUR file(s) (optional)")
    p.add_argument("--profile", default=None, help="AWS CLI profile to use from your credentials file")
    p.add_argument("--export-actions", default=None, help="Path to write actionable recommendations CSV (resource,action,estimated_saving,notes)")
    return p.parse_args()


def aws_client(service_name, region_name=None):
    if boto3 is None:
        raise RuntimeError("boto3 not available; install requirements.txt")
    # Use session if set
    global _BOTO_SESSION
    if _BOTO_SESSION:
        return _BOTO_SESSION.client(service_name, region_name=region_name) if region_name else _BOTO_SESSION.client(service_name)
    return boto3.client(service_name, region_name=region_name) if region_name else boto3.client(service_name)


def get_boto_session(profile=None):
    """Return a boto3.Session using an optional profile."""
    if boto3 is None:
        raise RuntimeError("boto3 not available; install requirements.txt")
    if profile:
        try:
            return boto3.Session(profile_name=profile)
        except Exception as e:
            # Profile not found or other issue; fall back to default session
            print(f"Warning: could not load profile '{profile}' ({e}). Falling back to default credentials.", file=sys.stderr)
            return boto3.Session()
    return boto3.Session()


def detect_account_id(dry_run=False):
    """Return AWS account id via STS. In dry-run, return a placeholder."""
    if dry_run or boto3 is None:
        return "000000000000"
    try:
        sts = aws_client("sts")
        resp = sts.get_caller_identity()
        return resp.get("Account")
    except Exception:
        return "unknown"


def fetch_costs_ce(days_back=180, dry_run=False, tag_keys=None):
    """Fetch cost and usage by service for the past months via Cost Explorer.

    Returns a list of dicts: {date: YYYY-MM-01, amount: float, service: str}
    """
    if dry_run:
        # Return service-level costs and an empty resource-level list for dry-run
        return SAMPLE_DATA["costs"], SAMPLE_DATA.get("resource_level", []), {}

    ce = aws_client("ce")
    end = datetime.date.today().replace(day=1)
    start = (end - datetime.timedelta(days=days_back)).replace(day=1)
    # Cost Explorer expects strings
    start_str = start.strftime("%Y-%m-%d")
    end_str = end.strftime("%Y-%m-%d")
    try:
        resp = ce.get_cost_and_usage(
            TimePeriod={"Start": start_str, "End": end_str},
            Granularity="MONTHLY",
            Metrics=["UnblendedCost"],
            GroupBy=[{"Type": "DIMENSION", "Key": "SERVICE"}],
        )
    except (BotoCoreError, ClientError) as e:
        print("Cost Explorer query failed:", e, file=sys.stderr)
        return []

    results = []
    for r in resp.get("ResultsByTime", []):
        date = r.get("TimePeriod", {}).get("Start")
        for g in r.get("Groups", []):
            service = g.get("Keys", ["Unknown"])[0]
            amount = float(g.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", 0.0))
            results.append({"date": date, "amount": amount, "service": service})
    # Try also to get resource-level costs when supported
    resource_level = []
    try:
        resp_r = ce.get_cost_and_usage(
            TimePeriod={"Start": start_str, "End": end_str},
            Granularity="MONTHLY",
            Metrics=["UnblendedCost"],
            GroupBy=[{"Type": "DIMENSION", "Key": "RESOURCE_ID"}],
        )
        for r in resp_r.get("ResultsByTime", []):
            date = r.get("TimePeriod", {}).get("Start")
            for g in r.get("Groups", []):
                rid = g.get("Keys", ["Unknown"])[0]
                amount = float(g.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", 0.0))
                resource_level.append({"date": date, "resource_id": rid, "amount": amount})
    except Exception:
        # Not all accounts/supports RESOURCE_ID grouping; that's fine
        resource_level = []

    # Tag-based aggregation
    tag_aggregates = {}
    if tag_keys:
        for tag in tag_keys:
            try:
                resp_t = ce.get_cost_and_usage(
                    TimePeriod={"Start": start_str, "End": end_str},
                    Granularity="MONTHLY",
                    Metrics=["UnblendedCost"],
                    GroupBy=[{"Type": "TAG", "Key": f"user:{tag}"}],
                )
                # collect latest month totals
                tag_totals = {}
                for r in resp_t.get("ResultsByTime", []):
                    for g in r.get("Groups", []):
                        k = g.get("Keys", ["Unknown"])[0]
                        amt = float(g.get("Metrics", {}).get("UnblendedCost", {}).get("Amount", 0.0))
                        tag_totals[k] = tag_totals.get(k, 0.0) + amt
                tag_aggregates[tag] = tag_totals
            except Exception:
                tag_aggregates[tag] = {}

    return results, resource_level, tag_aggregates


def parse_cur_local(path):
    """Parse a local CUR CSV (optionally gz) to extract per-resource costs.

    Returns a mapping resource_id -> total cost (USD) for the most recent month present in the file.
    """
    import gzip

    opener = gzip.open if path.endswith(".gz") else open
    data = {}
    try:
        with opener(path, "rt", encoding="utf-8") as f:
            # Read CSV into pandas for flexibility
            df = pd.read_csv(f)
    except Exception as e:
        print("Failed to read CUR file:", e, file=sys.stderr)
        return {}

    # Try to find resource id column
    candidate_cols = [c for c in df.columns if c.lower() in ("resourceid", "resource_id", "resource")]
    if not candidate_cols:
        # No resource id column; return empty
        return {}
    rid_col = candidate_cols[0]

    # cost column candidates
    cost_cols = [c for c in df.columns if c.lower() in ("unblendedcost", "cost", "lineitemunblendedcost")]
    if not cost_cols:
        return {}
    cost_col = cost_cols[0]

    # get most recent month
    date_cols = [c for c in df.columns if "date" in c.lower()]
    if date_cols:
        # attempt to use a date column
        try:
            df["_parsed_date"] = pd.to_datetime(df[date_cols[0]], errors="coerce")
            recent_month = df["_parsed_date"].dt.to_period("M").max()
            df_recent = df[df["_parsed_date"].dt.to_period("M") == recent_month]
        except Exception:
            df_recent = df
    else:
        df_recent = df

    grouped = df_recent.groupby(rid_col)[cost_col].sum()
    return grouped.to_dict()


def parse_cur_s3(bucket, prefix):
    """List and parse CUR files in S3 under bucket/prefix and aggregate resource costs. Returns resource->cost."""
    s3 = aws_client("s3")
    costs = {}
    try:
        resp = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)
        for obj in resp.get("Contents", []):
            key = obj.get("Key")
            # stream object
            tmp = "/tmp/cur_temp.csv"
            s3.download_file(bucket, key, tmp)
            part = parse_cur_local(tmp)
            for k, v in part.items():
                costs[k] = costs.get(k, 0.0) + v
    except Exception as e:
        print("Failed to read CUR from S3:", e, file=sys.stderr)
    return costs


def detect_wasted_resources(resources):
    """Detect wasted resources like unattached EBS volumes and unassociated Elastic IPs.

    Returns a dict with lists of wasted resource IDs.
    """
    wasted = {}
    # Unattached EBS volumes: EBS volumes list and describe to check attachments
    try:
        session = get_session()
        region = (session.region_name if session else None) or os.environ.get("AWS_REGION") or "us-east-1"
        ec2 = aws_client("ec2", region_name=region)
        vols = ec2.describe_volumes().get("Volumes", [])
        unattached = [v.get("VolumeId") for v in vols if not v.get("Attachments")]
        wasted["unattached_ebs"] = unattached
    except Exception:
        wasted["unattached_ebs"] = []

    # Unassociated Elastic IPs (not attached to an instance or network interface)
    try:
        ec2 = aws_client("ec2", region_name=region)
        addrs = ec2.describe_addresses().get("Addresses", [])
        unassoc = [a.get("AllocationId") or a.get("PublicIp") for a in addrs if not a.get("InstanceId") and not a.get("NetworkInterfaceId")]
        wasted["unassociated_eips"] = unassoc
    except Exception:
        wasted["unassociated_eips"] = []

    return wasted


def detect_wasted_resources_regions(dry_run=False, regions=None, concurrency=6):
    """Detect wasted resources across multiple regions in parallel.

    Returns a dict with lists of wasted resource IDs keyed by category.
    Categories: 'unattached_ebs', 'unassociated_eips'
    """
    if dry_run:
        return {"unattached_ebs": [], "unassociated_eips": []}

    session = get_session()
    # Determine regions to use
    if regions:
        region_list = regions
    else:
        # default to all available EC2 regions
        try:
            region_list = session.get_available_regions("ec2")
        except Exception:
            region_list = [session.region_name or os.environ.get("AWS_REGION") or "us-east-1"]

    from concurrent.futures import ThreadPoolExecutor, as_completed

    unattached = []
    unassoc_eips = []

    def check_region(region):
        r_unattached = []
        r_unassoc = []
        try:
            ec2 = session.client("ec2", region_name=region) if session else aws_client("ec2", region_name=region)
            # EBS volumes
            for page in ec2.get_paginator("describe_volumes").paginate():
                for v in page.get("Volumes", []):
                    if not v.get("Attachments"):
                        r_unattached.append(v.get("VolumeId"))
            # EIPs
            for page in ec2.get_paginator("describe_addresses").paginate():
                for a in page.get("Addresses", []):
                    if not a.get("InstanceId") and not a.get("NetworkInterfaceId"):
                        # prefer allocation id when present
                        r_unassoc.append(a.get("AllocationId") or a.get("PublicIp"))
        except Exception:
            pass
        return r_unattached, r_unassoc

    with ThreadPoolExecutor(max_workers=concurrency) as ex:
        futures = {ex.submit(check_region, r): r for r in region_list}
        for fut in as_completed(futures):
            try:
                r_unatt, r_unassoc = fut.result()
                unattached.extend(r_unatt)
                unassoc_eips.extend(r_unassoc)
            except Exception:
                pass

    return {"unattached_ebs": unattached, "unassociated_eips": unassoc_eips}


def inventory_resources(dry_run=False):
    # Legacy single-region inventory kept for backward compatibility
    return inventory_resources_regions(dry_run=dry_run, regions=None, concurrency=6)


def inventory_resources_regions(dry_run=False, regions=None, concurrency=6):
    """Collect inventory possibly across multiple regions. If regions is None, collect only global/one-shot resources.

    Returns a dict with counts and lists per resource.
    """
    if dry_run:
        merged = dict(SAMPLE_DATA["resources"])
        merged.update(SAMPLE_DATA.get("resource_lists", {}))
        return merged

    # Determine list of regions to check
    session = boto3.Session()
    if regions:
        region_list = regions
    else:
        # minimal default: use one region for global calls, but support --all-regions elsewhere
        region_list = []

    out = {}

    # Non-regional/global resources collected once
    try:
        s3 = session.client("s3")
        resp = s3.list_buckets()
        buckets = resp.get("Buckets", [])
        out["S3_buckets"] = len(buckets)
        out["S3_buckets_list"] = [b.get("Name") for b in buckets]
    except Exception:
        out["S3_buckets"] = "error"
        out["S3_buckets_list"] = []

    # Regional resources: we'll collect in parallel across requested regions
    # If region_list empty, default to single region from session or environment
    if not region_list:
        try:
            r = session.region_name or os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION")
            region_list = [r] if r else ["us-east-1"]
        except Exception:
            region_list = ["us-east-1"]

    from concurrent.futures import ThreadPoolExecutor, as_completed

    def collect_region(region):
        # Each region will return a dict of regional counts/lists
        d = {}
        try:
            ec2 = session.client("ec2", region_name=region)
            resp = ec2.describe_instances()
            ids = []
            for r in resp.get("Reservations", []):
                for inst in r.get("Instances", []):
                    ids.append(inst.get("InstanceId"))
            d["EC2_list"] = ids
            d["EC2"] = len(ids)
        except Exception:
            d["EC2"] = "error"
            d["EC2_list"] = []

        try:
            ebs = session.client("ec2", region_name=region).describe_volumes()
            v = [vv.get("VolumeId") for vv in ebs.get("Volumes", [])]
            d["EBS_volumes_list"] = v
            d["EBS_volumes"] = len(v)
        except Exception:
            d["EBS_volumes"] = "error"
            d["EBS_volumes_list"] = []

        try:
            lam = session.client("lambda", region_name=region)
            funcs = lam.list_functions()
            fs = [f.get("FunctionArn") for f in funcs.get("Functions", [])]
            d["Lambda_functions"] = len(fs)
            d["Lambda_functions_list"] = fs
        except Exception:
            d["Lambda_functions"] = "error"
            d["Lambda_functions_list"] = []

        try:
            rds = session.client("rds", region_name=region)
            dbs = rds.describe_db_instances().get("DBInstances", [])
            d["RDS"] = len(dbs)
            d["RDS_list"] = [db.get("DBInstanceIdentifier") for db in dbs]
        except Exception:
            d["RDS"] = "error"
            d["RDS_list"] = []

        try:
            ecr = session.client("ecr", region_name=region)
            repos = ecr.describe_repositories().get("repositories", [])
            d["ECR_repos"] = len(repos)
            d["ECR_repos_list"] = [r.get("repositoryName") for r in repos]
        except Exception:
            d["ECR_repos"] = "error"
            d["ECR_repos_list"] = []

        try:
            eks = session.client("eks", region_name=region)
            clusters = eks.list_clusters().get("clusters", [])
            d["EKS_clusters"] = len(clusters)
            d["EKS_clusters_list"] = clusters
        except Exception:
            d["EKS_clusters"] = "error"
            d["EKS_clusters_list"] = []

        try:
            sqs = session.client("sqs", region_name=region)
            q = sqs.list_queues().get("QueueUrls", []) or []
            d["SQS_queues"] = len(q)
            d["SQS_queues_list"] = q
        except Exception:
            d["SQS_queues"] = "error"
            d["SQS_queues_list"] = []

        try:
            efs = session.client("efs", region_name=region)
            fs = efs.describe_file_systems().get("FileSystems", [])
            d["EFS_file_systems"] = len(fs)
            d["EFS_file_systems_list"] = [f.get("FileSystemId") for f in fs]
        except Exception:
            d["EFS_file_systems"] = "error"
            d["EFS_file_systems_list"] = []

        return d

    aggregated = {}
    with ThreadPoolExecutor(max_workers=concurrency) as ex:
        futures = {ex.submit(collect_region, r): r for r in region_list}
        for fut in as_completed(futures):
            r = futures[fut]
            try:
                res = fut.result()
                for k, v in res.items():
                    # merge lists
                    if k.endswith("_list"):
                        aggregated.setdefault(k, [])
                        aggregated[k].extend(v)
                    else:
                        aggregated[k] = aggregated.get(k, 0) + (v if isinstance(v, int) else 0)
            except Exception:
                pass

    # Merge aggregated regional results into out
    out.update(aggregated)

    # Some additional global resources
    try:
        elbv2 = session.client("elbv2", region_name=region_list[0])
        lbs = elbv2.describe_load_balancers().get("LoadBalancers", [])
        out["LoadBalancers"] = len(lbs)
        out["LoadBalancers_list"] = [l.get("LoadBalancerArn") for l in lbs]
    except Exception:
        out["LoadBalancers"] = "error"
        out["LoadBalancers_list"] = []

    try:
        cf = session.client("cloudfront")
        dists = cf.list_distributions().get("DistributionList", {}).get("Items", [])
        out["CloudFront_distributions"] = len(dists)
        out["CloudFront_distributions_list"] = [i.get("Id") for i in dists]
    except Exception:
        out["CloudFront_distributions"] = "error"
        out["CloudFront_distributions_list"] = []

    try:
        ddb = session.client("dynamodb", region_name=region_list[0])
        tables = ddb.list_tables().get("TableNames", [])
        out["DynamoDB_tables"] = len(tables)
        out["DynamoDB_tables_list"] = tables
    except Exception:
        out["DynamoDB_tables"] = "error"
        out["DynamoDB_tables_list"] = []

    # AutoScaling (regional)
    try:
        asg = session.client("autoscaling", region_name=region_list[0])
        groups = asg.describe_auto_scaling_groups().get("AutoScalingGroups", [])
        out["AutoScalingGroups"] = len(groups)
        out["AutoScalingGroups_list"] = [g.get("AutoScalingGroupName") for g in groups]
    except Exception:
        out["AutoScalingGroups"] = "error"
        out["AutoScalingGroups_list"] = []

    return out

    out = {}
    try:
        ec2 = aws_client("ec2")
        resp = ec2.describe_instances()
        instances = 0
        ids = []
        for r in resp.get("Reservations", []):
            for inst in r.get("Instances", []):
                instances += 1
                ids.append(inst.get("InstanceId"))
        out["EC2"] = instances
        out["EC2_list"] = ids
    except Exception:
        out["EC2"] = "error"
        out["EC2_list"] = []

    try:
        rds = aws_client("rds")
        resp = rds.describe_db_instances()
        dbs = resp.get("DBInstances", [])
        out["RDS"] = len(dbs)
        out["RDS_list"] = [d.get("DBInstanceIdentifier") for d in dbs]
    except Exception:
        out["RDS"] = "error"
        out["RDS_list"] = []

    try:
        s3 = aws_client("s3")
        resp = s3.list_buckets()
        buckets = resp.get("Buckets", [])
        out["S3_buckets"] = len(buckets)
        out["S3_buckets_list"] = [b.get("Name") for b in buckets]
    except Exception:
        out["S3_buckets"] = "error"
        out["S3_buckets_list"] = []

    # Expanded inventory for common services
    # EBS volumes
    try:
        ec2 = aws_client("ec2")
        vols = ec2.describe_volumes()
        v = vols.get("Volumes", [])
        out["EBS_volumes"] = len(v)
        out["EBS_volumes_list"] = [vv.get("VolumeId") for vv in v]
    except Exception:
        out["EBS_volumes"] = "error"
        out["EBS_volumes_list"] = []

    # Elastic IPs
    try:
        addresses = ec2.describe_addresses()
        addrs = addresses.get("Addresses", [])
        out["EIPs"] = len(addrs)
        out["EIPs_list"] = [a.get("PublicIp") or a.get("AllocationId") for a in addrs]
    except Exception:
        out["EIPs"] = "error"
        out["EIPs_list"] = []

    # Load Balancers (ELBv2 / ALB/NLB)
    try:
        elbv2 = aws_client("elbv2")
        lbs = elbv2.describe_load_balancers()
        ll = lbs.get("LoadBalancers", [])
        out["LoadBalancers"] = len(ll)
        out["LoadBalancers_list"] = [l.get("LoadBalancerArn") for l in ll]
    except Exception:
        out["LoadBalancers"] = "error"
        out["LoadBalancers_list"] = []

    # Lambda functions
    try:
        lam = aws_client("lambda")
        funcs = lam.list_functions()
        fs = funcs.get("Functions", [])
        out["Lambda_functions"] = len(fs)
        out["Lambda_functions_list"] = [f.get("FunctionArn") for f in fs]
    except Exception:
        out["Lambda_functions"] = "error"
        out["Lambda_functions_list"] = []

    # CloudFront distributions
    try:
        cf = aws_client("cloudfront")
        dists = cf.list_distributions()
        items = (dists.get("DistributionList") or {}).get("Items", [])
        out["CloudFront_distributions"] = len(items)
        out["CloudFront_distributions_list"] = [i.get("Id") for i in items]
    except Exception:
        out["CloudFront_distributions"] = "error"
        out["CloudFront_distributions_list"] = []

    # DynamoDB tables
    try:
        ddb = aws_client("dynamodb")
        tables = ddb.list_tables()
        tbls = tables.get("TableNames", [])
        out["DynamoDB_tables"] = len(tbls)
        out["DynamoDB_tables_list"] = tbls
    except Exception:
        out["DynamoDB_tables"] = "error"
        out["DynamoDB_tables_list"] = []

    # EFS file systems
    try:
        efs = aws_client("efs")
        fss = efs.describe_file_systems()
        fslist = fss.get("FileSystems", [])
        out["EFS_file_systems"] = len(fslist)
        out["EFS_file_systems_list"] = [f.get("FileSystemId") for f in fslist]
    except Exception:
        out["EFS_file_systems"] = "error"
        out["EFS_file_systems_list"] = []

    # Auto Scaling Groups
    try:
        asg = aws_client("autoscaling")
        groups = asg.describe_auto_scaling_groups()
        ag = groups.get("AutoScalingGroups", [])
        out["AutoScalingGroups"] = len(ag)
        out["AutoScalingGroups_list"] = [g.get("AutoScalingGroupName") for g in ag]
    except Exception:
        out["AutoScalingGroups"] = "error"
        out["AutoScalingGroups_list"] = []

    # ECR repositories
    try:
        ecr = aws_client("ecr")
        repos = ecr.describe_repositories()
        rr = repos.get("repositories", [])
        out["ECR_repos"] = len(rr)
        out["ECR_repos_list"] = [r.get("repositoryName") for r in rr]
    except Exception:
        out["ECR_repos"] = "error"
        out["ECR_repos_list"] = []

    # EKS clusters
    try:
        eks = aws_client("eks")
        clusters = eks.list_clusters()
        cl = clusters.get("clusters", [])
        out["EKS_clusters"] = len(cl)
        out["EKS_clusters_list"] = cl
    except Exception:
        out["EKS_clusters"] = "error"
        out["EKS_clusters_list"] = []

    # SQS queues
    try:
        sqs = aws_client("sqs")
        queues = sqs.list_queues()
        ql = queues.get("QueueUrls", []) or []
        out["SQS_queues"] = len(ql)
        out["SQS_queues_list"] = ql
    except Exception:
        out["SQS_queues"] = "error"
        out["SQS_queues_list"] = []

    # SES identities
    try:
        ses = aws_client("ses")
        ids = ses.list_identities()
        il = ids.get("Identities", [])
        out["SES_identities"] = len(il)
        out["SES_identities_list"] = il
    except Exception:
        out["SES_identities"] = "error"
        out["SES_identities_list"] = []

    return out


def prepare_timeseries(cost_events):
    df = pd.DataFrame(cost_events)
    if df.empty:
        return pd.DataFrame()
    df["date"] = pd.to_datetime(df["date"]).dt.to_period("M").dt.to_timestamp()
    pivot = df.pivot_table(index="date", columns="service", values="amount", aggfunc="sum", fill_value=0.0)
    pivot["total"] = pivot.sum(axis=1)
    pivot = pivot.sort_index()
    return pivot


def forecast_linear(series, months=3):
    # Simple linear regression on month index
    if series.empty:
        return pd.Series(dtype=float)
    x = np.arange(len(series))
    y = series.values
    if len(x) < 2:
        # Not enough data: repeat last value
        last = y[-1] if len(y) else 0.0
        idx = pd.date_range(series.index[-1] + pd.offsets.MonthBegin(), periods=months, freq="MS")
        return pd.Series([last] * months, index=idx)
    A = np.vstack([x, np.ones_like(x)]).T
    m, c = np.linalg.lstsq(A, y, rcond=None)[0]
    future_x = np.arange(len(series), len(series) + months)
    preds = m * future_x + c
    idx = pd.date_range(series.index[-1] + pd.offsets.MonthBegin(), periods=months, freq="MS")
    return pd.Series(preds, index=idx)


def forecast_exponential_smoothing(series, months=3, alpha=0.3):
    """Simple exponential smoothing forecast as a companion to linear forecast.

    Returns a pd.Series of length `months` with future monthly estimates.
    """
    if series.empty:
        return pd.Series(dtype=float)
    # Use simple exponential smoothing to estimate next value
    values = series.astype(float).values
    s = values[0]
    for v in values[1:]:
        s = alpha * v + (1 - alpha) * s
    # s is the smoothed last value; project flat with small growth derived from recent slope
    if len(values) >= 2:
        recent_growth = (values[-1] - values[-2]) / max(values[-2], 1e-6)
    else:
        recent_growth = 0.0
    preds = []
    curr = s
    for m in range(months):
        curr = curr * (1 + recent_growth)
        preds.append(curr)
    idx = pd.date_range(series.index[-1] + pd.offsets.MonthBegin(), periods=months, freq="MS")
    return pd.Series(preds, index=idx)


REPORT_TEMPLATE = """
<html>
<head>
    <meta charset="utf-8" />
    <title>AWS Cost Analysis Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 24px; color: #222; }
        header { display:flex; justify-content:space-between; align-items:center }
        .metrics { display:flex; gap:12px; margin-top:12px }
        .metric { background:#f5f7fb; padding:12px; border-radius:6px; box-shadow:0 1px 2px rgba(0,0,0,0.05) }
        .charts { display:flex; flex-wrap:wrap; gap:12px; margin-top:12px }
        .chart { background:white; padding:8px; border:1px solid #eee }
        h1 { margin:0 }
        footer { margin-top:32px; border-top:1px solid #ddd; padding-top:8px; color:#666 }
        .definitions { background:#fffaf0; padding:12px; border-radius:6px }
    </style>
</head>
<body>
<header>
    <div>
        <h1>AWS Cost Analysis</h1>
    <div>AWS Cost Analysis Tool by <a href="https://www.linkedin.com/in/jaydenshutt/">Jayden Shutt</a></div>
    </div>
    <div style="text-align:right">
        <div>Analysis period: <strong>{{ start }} — {{ end }}</strong></div>
        <div>Total monthly cost: <strong>${{ '%.2f'|format(latest_total) }}</strong></div>
        <div>Account: <strong>{{ account_id }}</strong></div>
    </div>
</header>

<section class="metrics" aria-label="summary-metrics">
    <div class="metric"><strong>Latest total</strong><div>${{ '%.2f'|format(latest_total) }}</div></div>
    <div class="metric"><strong>Top service</strong><div>{% if top_services %}{{ top_services[0][0] }} — ${{ '%.2f'|format(top_services[0][1]) }}{% else %}N/A{% endif %}</div></div>
    <div class="metric"><strong>Resources inspected</strong><div>{{ resources|length }}</div></div>
</section>

<section>
    <h2>Executive summary</h2>
    <p>{{ executive_summary }}</p>
    <p>This document summarizes monthly spending trends, identifies top-cost services and resources, and provides prioritized, prescriptive recommendations written in plain language for non-technical executives. Charts below visualize historical spend, service breakdown, and short-term forecast.</p>
</section>

<section class="charts">
    <div class="chart"><h3>Trend</h3><img src="{{ cost_plot }}" alt="Total cost trend" style="max-width:100%;"></div>
    <div class="chart"><h3>Service share (latest)</h3><img src="{{ service_pie }}" alt="Service share" style="max-width:100%; height:220px"></div>
    <div class="chart"><h3>Stacked services</h3><img src="{{ stacked_area }}" alt="Stacked area" style="max-width:100%;"></div>
    <div class="chart"><h3>Top services (bar)</h3><img src="{{ top_bar }}" alt="Top services bar" style="max-width:100%; height:220px"></div>
    <div class="chart"><h3>Forecast</h3><img src="{{ forecast_plot }}" alt="Forecast" style="max-width:100%;"></div>
</section>
<h3>Top services</h3>
<ul>
{% for svc, amt in top_services %}
  <li>{{ svc }}: ${{ '%.2f'|format(amt) }}</li>
{% endfor %}
</ul>
<h3>Recommendations</h3>
<ul>
{% for r in recommendations %}
  <li>{{ r }}</li>
{% endfor %}
</ul>
<h4>Estimated monthly savings (by recommendation)</h4>
{% if rec_objects %}
<ul>
{% for o in rec_objects %}
    <li><strong>{{ o.service }}</strong> — {{ o.note }}: Estimated savings: ${{ '%.2f'|format(o.estimated_monthly_savings) }}</li>
{% endfor %}
</ul>
{% else %}
<p>No estimated savings available.</p>
{% endif %}
<h3>Prescriptive guidance (by service)</h3>
{% for svc, info in per_service_advice.items() %}
    <section style="background:#f9f9fb;padding:10px;border-radius:6px;margin-bottom:8px">
        <h4>{{ svc }}</h4>
        <p><em>{{ info.explanation }}</em></p>
        <p><strong>Recommended actions:</strong></p>
        <ul>
        {% for a in info.actions %}
            <li>{{ a }}</li>
        {% endfor %}
        </ul>
        <p><strong>Estimated monthly saving (conservative):</strong> ${{ '%.2f'|format(info.estimated_monthly_savings) }}</p>
    </section>
{% endfor %}

<h3>Glossary (plain language)</h3>
<dl>
{% for term, desc in glossary.items() %}
    <dt><strong>{{ term }}</strong></dt>
    <dd>{{ desc }}</dd>
{% endfor %}
</dl>
<h3>Top resources (by cost)</h3>
<ul>
{% for rid, amt in top_resources %}
    <li>{{ rid }}: ${{ '%.2f'|format(amt) }}</li>
{% endfor %}
</ul>
<h3>Tag aggregates</h3>
{% for tag, mapping in tag_aggregates.items() %}
    <h4>Tag: {{ tag }}</h4>
    <ul>
    {% for tv, v in mapping.items() %}
        <li>{{ tv }}: ${{ '%.2f'|format(v) }}</li>
    {% endfor %}
    </ul>
{% endfor %}
<h3>Wasted resources (potential)</h3>
<ul>
{% for k, lst in wasted_resources.items() %}
    <li>{{ k }}: {{ lst|length }} items</li>
    <ul>
    {% for item in lst[:20] %}
        <li>{{ item }}</li>
    {% endfor %}
    </ul>
{% endfor %}
</ul>
<h2>Graphs</h2>
<img src="{{ cost_plot }}" alt="costs">
<img src="{{ service_pie }}" alt="pie">
<h2>Resource inventory</h2>
<pre>{{ resources }}</pre>

</body>
</html>
"""


def generate_plots(pivot, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    cost_plot = os.path.join(out_dir, "costs.png")
    service_pie = os.path.join(out_dir, "top_services.png")
    stacked_area = os.path.join(out_dir, "stacked_area.png")
    top_bar = os.path.join(out_dir, "top_services_bar.png")
    forecast_plot = os.path.join(out_dir, "forecast.png")

    if pivot.empty:
        # create placeholder
        fig = plt.figure(figsize=(6, 3))
        plt.text(0.5, 0.5, "No data", ha="center", va="center")
        fig.savefig(cost_plot)
        plt.close(fig)
        fig = plt.figure(figsize=(4, 4))
        plt.text(0.5, 0.5, "No data", ha="center", va="center")
        fig.savefig(service_pie)
        plt.close(fig)
        return cost_plot, service_pie

    fig, ax = plt.subplots(figsize=(10, 4))
    pivot["total"].plot(ax=ax, marker="o")
    ax.set_title("Total monthly cost")
    ax.set_ylabel("USD")
    fig.tight_layout()
    fig.savefig(cost_plot)
    plt.close(fig)

    # Stacked area of services
    services_only = pivot.drop(columns=["total"]) if "total" in pivot.columns else pivot
    if not services_only.empty:
        fig3, ax3 = plt.subplots(figsize=(10, 4))
        services_only.plot(kind="area", stacked=True, ax=ax3)
        ax3.set_title("Monthly cost by service (stacked)")
        ax3.set_ylabel("USD")
        fig3.tight_layout()
        fig3.savefig(stacked_area)
        plt.close(fig3)
    else:
        # placeholder
        fig3 = plt.figure(figsize=(10, 4))
        plt.text(0.5, 0.5, "No data", ha="center")
        fig3.savefig(stacked_area)
        plt.close(fig3)

    # Pie of last month by service
    last = pivot.iloc[-1].drop(labels=["total"]).sort_values(ascending=False)
    fig2, ax2 = plt.subplots(figsize=(6, 6))
    if last.sum() > 0:
        ax2.pie(last, labels=last.index, autopct="%.1f%%")
        ax2.set_title("Cost share by service (latest)")
    else:
        ax2.text(0.5, 0.5, "No cost data", ha="center", va="center")
    fig2.tight_layout()
    fig2.savefig(service_pie)
    plt.close(fig2)

    # Bar chart of top services
    topn = last.head(10)
    fig4, ax4 = plt.subplots(figsize=(8, 4))
    topn.plot(kind="bar", ax=ax4)
    ax4.set_title("Top services (latest month)")
    ax4.set_ylabel("USD")
    fig4.tight_layout()
    fig4.savefig(top_bar)
    plt.close(fig4)

    # Forecast overlay (if we can compute a simple linear forecast)
    try:
        from pandas import Series
        series = pivot["total"]
        if not series.empty:
            months = 3
            preds = forecast_linear(series, months=months)
            fig5, ax5 = plt.subplots(figsize=(10, 4))
            series.plot(ax=ax5, marker="o", label="Historical")
            preds.plot(ax=ax5, marker="x", linestyle="--", label="Forecast")
            ax5.set_title("Historical and forecasted total cost")
            ax5.set_ylabel("USD")
            ax5.legend()
            fig5.tight_layout()
            fig5.savefig(forecast_plot)
            plt.close(fig5)
        else:
            # placeholder
            fig5 = plt.figure(figsize=(10, 4))
            plt.text(0.5, 0.5, "No forecast data", ha="center")
            fig5.savefig(forecast_plot)
            plt.close(fig5)
    except Exception:
        pass

    return {
        "cost_plot": cost_plot,
        "service_pie": service_pie,
        "stacked_area": stacked_area,
        "top_bar": top_bar,
        "forecast_plot": forecast_plot,
    }


def analyze(args):
    tag_keys = [k.strip() for k in args.tag_keys.split(",")] if args.tag_keys else None
    costs, resource_level_costs, tag_aggregates = fetch_costs_ce(days_back=args.months * 31, dry_run=args.dry_run, tag_keys=tag_keys)
    resources = inventory_resources(dry_run=args.dry_run)

    pivot = prepare_timeseries(costs)

    # Ensure all outputs go into an output/ subfolder next to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_output_dir = os.path.join(script_dir, "output")
    os.makedirs(base_output_dir, exist_ok=True)

    # Determine target HTML, PDF, and CSV/XLSX paths inside output/ (use provided basenames)
    html_path = os.path.join(base_output_dir, os.path.basename(args.output)) if args.output else os.path.join(base_output_dir, "report.html")
    pdf_path = os.path.join(base_output_dir, os.path.basename(args.pdf)) if args.pdf else None
    csv_export_path = None
    if args.csv:
        csv_export_path = os.path.join(base_output_dir, os.path.basename(args.csv))
        pivot.to_csv(csv_export_path)

    # Forecast
    forecast_months = args.forecast_months
    forecast = forecast_linear(pivot["total"] if "total" in pivot else pd.Series(dtype=float), months=forecast_months)
    # complementary exponential smoothing forecast to catch recent trend changes
    exp_forecast = forecast_exponential_smoothing(pivot["total"] if "total" in pivot else pd.Series(dtype=float), months=forecast_months)

    # Top services in latest month
    top_services = []
    latest_total = 0.0
    if not pivot.empty:
        latest = pivot.iloc[-1].drop(labels=["total"]) if "total" in pivot.columns else pivot.iloc[-1]
        latest_total = float(pivot["total"].iloc[-1]) if "total" in pivot.columns else float(latest.sum())
        top = latest.sort_values(ascending=False).head(5)
        top_services = [(svc, float(v)) for svc, v in top.items()]

    # Top resources if resource-level costs returned
    top_resources = []
    if resource_level_costs:
        df_res = pd.DataFrame(resource_level_costs)
        if not df_res.empty:
            df_res["date"] = pd.to_datetime(df_res["date"]).dt.to_period("M").dt.to_timestamp()
            last_date = df_res["date"].max()
            lastdf = df_res[df_res["date"] == last_date]
            topr = lastdf.groupby("resource_id").sum()["amount"].sort_values(ascending=False).head(10)
            top_resources = [(rid, float(v)) for rid, v in topr.items()]

        # CUR parsing if requested
        cur_resource_costs = {}
        if args.cur_file:
            cur_resource_costs = parse_cur_local(args.cur_file)
        elif args.cur_s3_bucket and args.cur_prefix:
            cur_resource_costs = parse_cur_s3(args.cur_s3_bucket, args.cur_prefix)

        # Merge CUR-derived resource costs if available (prefer CUR values for per-resource accuracy)
        if cur_resource_costs:
            cr_sorted = sorted(cur_resource_costs.items(), key=lambda x: x[1], reverse=True)[:20]
            top_resources = [(rid, float(v)) for rid, v in cr_sorted]
    # Detect wasted resources (best-effort)
    # If user requested multi-region scanning, use the multi-region detector for completeness
    if args.all_regions or args.regions:
        regions = None
        if args.regions:
            regions = [r.strip() for r in args.regions.split(",")]
        wasted = detect_wasted_resources_regions(dry_run=args.dry_run, regions=regions, concurrency=args.concurrency)
    else:
        wasted = detect_wasted_resources(resources)

    # Build richer, prescriptive recommendations (senior cloud architect style)
    def build_recommendations(latest_total, pivot, resources, top_services, wasted, cur_resource_costs, tag_aggregates):
        recs = []
        rec_objs = []
        per_service = {}

        # Friendly executive summary
        if latest_total and pivot.shape[0] >= 2:
            # determine recent growth rate (month-over-month)
            recent = pivot['total'].pct_change().dropna()
            avg_growth = float(recent.tail(3).mean()) if not recent.empty else 0.0
            trend = "increasing" if avg_growth > 0.03 else ("decreasing" if avg_growth < -0.03 else "stable")
            exec_summary = f"Your account's most recent monthly spend is ${latest_total:.2f}. Over the period analyzed the spend appears {trend} (recent average monthly change {avg_growth*100:.1f}%). The report highlights high-cost services and practical steps a cloud architect would prioritize to reduce recurring spend."
        else:
            exec_summary = f"Latest monthly spend is ${latest_total:.2f}. Insufficient historical data to determine trend confidently."

        # Service-specific templates
        def add_service_advice(service_label, explanation, actions, saving_factor=0.05):
            # saving_factor is proportion of latest_total as conservative estimate
            est = latest_total * saving_factor if latest_total else 0.0
            per_service[service_label] = {
                "explanation": explanation,
                "actions": actions,
                "estimated_monthly_savings": est,
            }
            recs.append(f"{service_label}: {explanation} Recommended actions: {', '.join(actions)}. Estimated potential monthly saving: ${est:.2f}.")
            rec_objs.append({"id": f"rec_{service_label}", "service": service_label, "estimated_monthly_savings": est, "note": "; ".join(actions)})

        # Generic service mappings
        add_service_advice(
            "AmazonEC2",
            "Virtual machine instances that power applications. Costs come from instance type, count, uptime, and attached storage.",
            ["rightsizing/resize instances", "use Savings Plans or Reserved Instances for steady-state workloads", "use Spot for fault-tolerant, non-prod workloads", "implement autoscaling to remove idle capacity"],
            saving_factor=0.15,
        )
        add_service_advice(
            "AmazonRDS",
            "Managed relational databases. Costs depend on instance size, provisioned IOPS, and backup/retention settings.",
            ["identify idle or low-utilization DBs and downsize", "consider serverless or Aurora where suitable", "use reserved instances for steady workloads"],
            saving_factor=0.10,
        )
        add_service_advice(
            "AmazonS3",
            "Object storage where costs come from stored data, class (Standard/IA/Glacier), and requests/egress.",
            ["apply lifecycle policies to move infrequently accessed data to IA/Glacier", "enable intelligent tiering", "clean up orphaned objects and old backups"],
            saving_factor=0.05,
        )
        add_service_advice(
            "AmazonEBS",
            "Block storage volumes attached to instances; unattached volumes still incur charges.",
            ["delete unattached volumes or snapshot and delete after validation", "use gp3 for cost-effective performance"],
            saving_factor=0.02,
        )
        add_service_advice(
            "AmazonLambda",
            "Serverless function invocations billed by duration and memory. Cost can grow with inefficient code or excessive concurrency.",
            ["review function memory/time settings and reduce over-provisioning", "optimize code paths to reduce execution time"],
            saving_factor=0.02,
        )
        add_service_advice(
            "AmazonECR",
            "Container image registry costs come from storage and data transfer.",
            ["clean up old unused images and enable lifecycle policies", "scan and delete large untagged images"],
            saving_factor=0.01,
        )
        add_service_advice(
            "AmazonEKS",
            "Kubernetes control planes are managed; cluster costs also include underlying EC2 nodes, storage, and managed add-ons.",
            ["rightsize node groups, use node auto-provisioners, consider Fargate for small workloads"],
            saving_factor=0.08,
        )
        add_service_advice(
            "AmazonDynamoDB",
            "NoSQL database billed by read/write capacity or on-demand usage and storage.",
            ["switch to on-demand if usage is spiky, or reduce provisioned capacity and use autoscaling", "clean up unused global secondary indexes (GSIs)"],
            saving_factor=0.03,
        )
        add_service_advice(
            "AmazonCloudFront",
            "CDN billed by data transfer and requests; optimization includes caching and compressing content.",
            ["increase cache TTLs, enable compression and caching headers, use regional edge caches smartly"],
            saving_factor=0.02,
        )
        add_service_advice(
            "ElasticLoadBalancing",
            "Load balancers incur hourly and data-processing charges.",
            ["remove idle load balancers, consolidate listeners, and use application-level routing when possible"],
            saving_factor=0.02,
        )

        # Wasted resources specifics
        if wasted:
            # unattached EBS
            unattached = wasted.get('unattached_ebs', [])
            if unattached:
                est = min(len(unattached) * 5.0, latest_total * 0.02 if latest_total else 0.0)
                recs.append(f"Found {len(unattached)} unattached EBS volumes. These typically cost a small fixed amount per month; consider snapshot+delete or delete after validation. Potential account-level saving: ${est:.2f}.")
                rec_objs.append({"id": "wasted_ebs", "service": "AmazonEBS", "estimated_monthly_savings": est, "note": "delete or snapshot unattached volumes"})
            unassoc = wasted.get('unassociated_eips', [])
            if unassoc:
                est = min(len(unassoc) * 3.0, latest_total * 0.01 if latest_total else 0.0)
                recs.append(f"Found {len(unassoc)} unassociated Elastic IPs. These carry a small monthly cost; consider releasing them. Potential saving: ${est:.2f}.")
                rec_objs.append({"id": "wasted_eip", "service": "EC2/EIP", "estimated_monthly_savings": est, "note": "release unassociated EIPs"})

        # Top-services targeted message override/additions
        if top_services:
            # top_services are tuples (service, amount)
            top_svc_name = top_services[0][0]
            # make a friendly name mapping
            svc_key = top_svc_name
            # assign extra tailored action if known
            if svc_key.lower().startswith('amazonec2'):
                recs.append("As EC2 is the largest cost, prioritize a rightsizing and purchase strategy: perform instance-level utilization analysis (CPU, memory, network) and purchase Savings Plans for stable workloads. This often yields the largest predictable savings.")
            elif svc_key.lower().startswith('amazons3'):
                recs.append("As S3 shows significant cost, prioritize lifecycle policies and inventory the largest buckets/objects; consider archival and deduplication of large backups.")

        # Glossary (short, executive-friendly)
        glossary = {
            "Monthly cost": "Total billed amount for the month (approx).",
            "Rightsizing": "Matching compute or database size to actual usage (reduces overprovisioning).",
            "Savings Plans / Reserved Instances": "Commitment-based discounts in exchange for 1-3 year usage commitments.",
            "Spot instances": "Deeply discounted compute for interruptible workloads.",
        }

        return recs, rec_objs, exec_summary, per_service, glossary

    # call build_recommendations
    recommendations, rec_objects, executive_summary, per_service_advice, glossary = build_recommendations(latest_total, pivot, resources, top_services, wasted, (locals().get('cur_resource_costs') or {}), tag_aggregates)

    # detect wasted resources (best-effort)
    # If user requested multi-region scanning, use the multi-region detector for completeness
    if args.all_regions or args.regions:
        regions = None
        if args.regions:
            regions = [r.strip() for r in args.regions.split(",")]
        wasted = detect_wasted_resources_regions(dry_run=args.dry_run, regions=regions, concurrency=args.concurrency)
    else:
        wasted = detect_wasted_resources(resources)

    # Render report
    out_dir = os.path.join(base_output_dir, "assets")
    plots = generate_plots(pivot, out_dir)
    # map plots to variables for PDF use
    cost_plot = plots.get("cost_plot")
    service_pie = plots.get("service_pie")
    stacked_area = plots.get("stacked_area")
    top_bar = plots.get("top_bar")
    forecast_plot = plots.get("forecast_plot")

    tpl = Template(REPORT_TEMPLATE)
    start = pivot.index[0].strftime("%Y-%m-%d") if not pivot.empty else "N/A"
    end = pivot.index[-1].strftime("%Y-%m-%d") if not pivot.empty else "N/A"
    # detect AWS account id (dry-run uses placeholder)
    account_id = detect_account_id(dry_run=args.dry_run)
    html = tpl.render(
        start=start,
        end=end,
        latest_total=latest_total,
        top_services=top_services,
        top_resources=top_resources,
        tag_aggregates=tag_aggregates,
        wasted_resources=wasted,
        recommendations=recommendations,
        rec_objects=rec_objects,
        executive_summary=executive_summary,
        per_service_advice=per_service_advice,
        glossary=glossary,
        account_id=account_id,
        cost_plot=os.path.relpath(cost_plot, os.path.dirname(html_path)),
        service_pie=os.path.relpath(service_pie, os.path.dirname(html_path)),
        stacked_area=os.path.relpath(stacked_area, os.path.dirname(html_path)),
        top_bar=os.path.relpath(top_bar, os.path.dirname(html_path)),
        forecast_plot=os.path.relpath(forecast_plot, os.path.dirname(html_path)),
        resources=json.dumps(resources, indent=2),
    )

    # Do not persist the HTML file; we render the template but only export PDF/CSV/XLSX
    # Keeping HTML rendering in-memory allows image generation and templating checks, but we won't write the HTML file to disk.

    # Optionally write a PDF summary
    if args.pdf:
        try:
            if not pdf_path:
                pdf_path = os.path.join(base_output_dir, os.path.basename(args.pdf))
            write_pdf_report(
                pdf_path,
                start,
                end,
                latest_total,
                top_services,
                recommendations,
                cost_plot,
                service_pie,
                resources,
                top_resources=top_resources,
                tag_aggregates=tag_aggregates,
                wasted=wasted,
                stacked_area=stacked_area,
                top_bar=top_bar,
                forecast_plot=forecast_plot,
                rec_objects=rec_objects,
                executive_summary=executive_summary,
                per_service_advice=per_service_advice,
                glossary=glossary,
                    actionable_rows=locals().get('actionable_rows', None),
                    account_id=account_id,
            )
            print(f"PDF written to {pdf_path}")
        except Exception as e:
            print("Failed to write PDF:", e, file=sys.stderr)
    # Ensure no HTML file is left on disk — we render HTML in-memory for templating but do not persist it
    try:
        if os.path.exists(html_path):
            # only remove if the path is inside our base_output_dir to avoid accidental deletions
            if os.path.commonpath([os.path.abspath(html_path), base_output_dir]) == os.path.abspath(base_output_dir):
                try:
                    os.remove(html_path)
                except Exception:
                    pass
    except Exception:
        pass
    if args.export_actions:
        import csv
        rows = []
        # rec_objects: higher-level recommendations
        def compute_priority_confidence(est_value, latest_total, source="heuristic", extra=None):
            """Compute priority label, confidence label, a granular confidence reason, and a numeric priority_score.

            Returns: (priority_label, confidence_label, confidence_reason, priority_score(float))
            priority_score is a numeric value useful for sorting (higher = more urgent). We compute it as
            est_value normalized by latest_total plus small boosts for high-confidence sources and wasted items.
            """
            pr = "Low"
            conf = "Medium"
            reason = "Heuristic estimate based on service-level spend"
            score = 0.0
            try:
                pct = (est_value / latest_total) if latest_total and latest_total > 0 else 0.0
                # base score: percentage of total scaled to 100
                score = pct * 100.0
                # ensure est_value absolute floor contributes
                score += min(est_value / 10.0, 10.0)

                # priority thresholds
                if pct >= 0.05 or est_value >= 100.0:
                    pr = "High"
                elif pct >= 0.01 or est_value >= 20.0:
                    pr = "Medium"
                else:
                    pr = "Low"

                # source-based confidence and reason augmentation
                if source == "wasted":
                    conf = "High"
                    # include counts when available in extra
                    if isinstance(extra, dict) and extra.get("count") is not None:
                        reason = f"Detected {extra.get('count')} unused resource(s) via inventory scan"
                        # boost the score modestly for multiple detected items
                        score += min(extra.get('count') * 2.0, 20.0)
                    else:
                        reason = "Detected unused resource via inventory scan"
                elif source == "cur":
                    conf = "High"
                    reason = "CUR provided: per-resource cost available"
                    score += 10.0
                elif source == "exact":
                    conf = "High"
                    reason = "Exact per-resource cost available (high confidence)"
                    score += 8.0
                else:
                    conf = "Medium"
                    # if caller provided contextual extra (e.g., rule), append
                    if extra:
                        reason = f"Heuristic: {extra}"
                    else:
                        reason = "Heuristic estimate based on service-level spend"
            except Exception:
                pr = "Low"
                conf = "Low"
                reason = "Failed to compute confidence"
                score = 0.0

            # clamp score
            try:
                score = float(score)
            except Exception:
                score = 0.0
            return pr, conf, reason, score

        for o in (rec_objects or []):
            est = float(o.get('estimated_monthly_savings', 0.0) or 0.0)
            pr, conf, reason, score = compute_priority_confidence(est, latest_total, source="heuristic", extra=o.get('id'))
            rows.append({
                "resource": o.get("service") or "account",
                "action": o.get("note") or "recommendation",
                "estimated_monthly_saving": f"{est:.2f}",
                "priority": pr,
                "priority_score": f"{score:.2f}",
                "confidence": conf,
                "confidence_reason": reason,
                "notes": "autogenerated recommendation",
            })
        # wasted resources as per-resource actions
        if wasted:
            for k, lst in wasted.items():
                for item in lst:
                    # conservative per-resource estimate for wasted items
                    est_val = 0.0
                    if "ebs" in k.lower():
                        est_val = 5.0
                    elif "eip" in k.lower():
                        est_val = 3.0
                    # pass count for better confidence_reason
                    pr, conf, reason, score = compute_priority_confidence(est_val, latest_total, source="wasted", extra={"count": len(lst)})
                    # make reason more granular by including detected counts
                    if conf == "High" and isinstance(reason, str) and "Detected" not in reason:
                        reason = f"Detected {len(lst)} {k} items via inventory scan"
                    rows.append({
                        "resource": str(item),
                        "action": f"review and/or delete ({k})",
                        "estimated_monthly_saving": f"{est_val:.2f}" if est_val else "",
                        "priority": pr,
                        "priority_score": f"{score:.2f}",
                        "confidence": conf,
                        "confidence_reason": reason,
                        "notes": f"wasted resource detected: {k}",
                    })

        # write CSV
            try:
                # if the user specified a directory in args.export_actions, use it; otherwise place CSV in base_output_dir
                if os.path.isabs(args.export_actions) or os.path.dirname(args.export_actions):
                    csv_path = args.export_actions
                else:
                    csv_path = os.path.join(base_output_dir, os.path.basename(args.export_actions))

                # ensure CSV path ends with .csv
                if not csv_path.lower().endswith('.csv'):
                    csv_path = csv_path + '.csv'

                os.makedirs(os.path.dirname(csv_path), exist_ok=True)

                with open(csv_path, "w", newline="", encoding="utf-8") as cf:
                    fieldnames = ["resource", "action", "estimated_monthly_saving", "priority", "priority_score", "confidence", "confidence_reason", "notes"]
                    writer = csv.DictWriter(cf, fieldnames=fieldnames)
                    writer.writeheader()
                    for r in rows:
                        # ensure all fields present
                        for fn in fieldnames:
                            if fn not in r:
                                r[fn] = ""
                        writer.writerow(r)
                print(f"Actionable recommendations exported to {csv_path}")

                # write XLSX alongside the CSV for easier operations team handoff and apply currency formatting when possible
                try:
                    xlsx_path = os.path.splitext(csv_path)[0] + '.xlsx'
                    df_rows = pd.DataFrame(rows)
                    # ensure column order
                    cols = ["resource", "action", "estimated_monthly_saving", "priority", "priority_score", "confidence", "confidence_reason", "notes"]
                    df_rows = df_rows.reindex(columns=cols)

                    # Try to write XLSX with currency formatting using openpyxl if available
                    try:
                        from openpyxl import load_workbook
                        from openpyxl.styles import numbers
                        # write initial dataframe to xlsx
                        df_rows.to_excel(xlsx_path, index=False)
                        # open and apply currency format to the estimated_monthly_saving column
                        wb = load_workbook(xlsx_path)
                        ws = wb.active
                        # find the column index for estimated_monthly_saving
                        header = [c.value for c in ws[1]]
                        if "estimated_monthly_saving" in header:
                            col_idx = header.index("estimated_monthly_saving") + 1
                            # apply currency number format to all data rows in that column
                            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                                for cell in row:
                                    # attempt to coerce to float if possible
                                    try:
                                        val = float(str(cell.value)) if cell.value is not None and str(cell.value) != "" else None
                                        if val is not None:
                                            cell.value = val
                                            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                                    except Exception:
                                        pass
                        wb.save(xlsx_path)
                        print(f"Actionable recommendations exported to {xlsx_path}")
                    except Exception:
                        # fallback simple write if openpyxl isn't available or formatting fails
                        df_rows.to_excel(xlsx_path, index=False)
                        print(f"Actionable recommendations exported to {xlsx_path} (unformatted) ")
                except Exception as e:
                    print("Failed to write XLSX alongside CSV:", e, file=sys.stderr)
            except Exception as e:
                print("Failed to write actionable CSV:", e, file=sys.stderr)

        # make rows available for PDF injection by attaching to a variable in outer scope
        actionable_rows = rows


def write_pdf_report(pdf_path, start, end, latest_total, top_services, recommendations, cost_plot, service_pie, resources, top_resources=None, tag_aggregates=None, wasted=None, stacked_area=None, top_bar=None, forecast_plot=None, rec_objects=None, executive_summary=None, per_service_advice=None, glossary=None, actionable_rows=None, account_id=None):
    if letter is None:
        raise RuntimeError("ReportLab not installed. Add reportlab to requirements and install it.")

    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Author attribution
    author_para = Paragraph("AWS Cost Analysis Tool by <a href=\"https://www.linkedin.com/in/jaydenshutt/\">Jayden Shutt</a>", styles["Normal"])
    story.append(author_para)
    story.append(Spacer(1, 6))

    story.append(Paragraph("AWS Account Cost Analysis", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Analysis period: {start} to {end}", styles["Normal"]))
    story.append(Paragraph(f"Latest total monthly cost: ${latest_total:.2f}", styles["Normal"]))
    if account_id:
        story.append(Paragraph(f"Account: {account_id}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Executive summary (if provided)
    if executive_summary:
        story.append(Paragraph("Executive summary", styles["Heading2"]))
        story.append(Paragraph(executive_summary, styles["Normal"]))
        story.append(Spacer(1, 12))

    # Top services table
    story.append(Paragraph("Top Services", styles["Heading2"]))
    data = [["Service", "Amount (USD)"]]
    for svc, amt in top_services:
        data.append([svc, f"${amt:.2f}"])
    t = Table(data, hAlign="LEFT")
    t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
    story.append(t)
    story.append(Spacer(1, 12))

    # Recommendations
    story.append(Paragraph("Recommendations", styles["Heading2"]))
    for r in recommendations:
        story.append(Paragraph(f"- {r}", styles["Normal"]))
    story.append(Spacer(1, 12))
    # Estimated savings table
    if rec_objects:
        story.append(Paragraph("Estimated savings (by recommendation)", styles["Heading2"]))
        # Use Paragraphs inside table cells so long notes wrap instead of overflowing the page
        from reportlab.lib.styles import ParagraphStyle
        small_para = ParagraphStyle("small_para", parent=styles["Normal"], fontSize=9, leading=11)
        header_style = ParagraphStyle("hdr", parent=styles["Normal"], fontSize=10, leading=12)

        data = [
            [Paragraph("<b>Service</b>", header_style), Paragraph("<b>Note</b>", header_style), Paragraph("<b>Est. monthly saving (USD)</b>", header_style)]
        ]
        for o in rec_objects:
            svc = Paragraph(str(o.get('service') or ''), small_para)
            note = Paragraph(str(o.get('note') or ''), small_para)
            amt = Paragraph(f"${o.get('estimated_monthly_savings', 0.0):.2f}", small_para)
            data.append([svc, note, amt])

        # set column widths so the note column wraps
        col_widths = [110, 320, 90]
        t = Table(data, colWidths=col_widths, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    # Actionable rows table (from CSV/XLSX export)
    if actionable_rows:
        story.append(Paragraph("Actionable items (sorted by priority_score)", styles["Heading2"]))
        # sort rows by numeric priority_score descending if present
        try:
            sorted_rows = sorted(actionable_rows, key=lambda r: float(r.get('priority_score') or 0.0), reverse=True)
        except Exception:
            sorted_rows = actionable_rows
        # limit rows to a reasonable number for PDF (e.g., 200)
        max_show = 200
        cols = ["Resource", "Action", "Est Saving (USD)", "Priority", "Priority Score", "Confidence", "Confidence Reason"]
        data = [cols]
        for r in sorted_rows[:max_show]:
            data.append([
                str(r.get('resource', '')),
                str(r.get('action', '')),
                str(r.get('estimated_monthly_saving', '')),
                str(r.get('priority', '')),
                str(r.get('priority_score', '')),
                str(r.get('confidence', '')),
                str(r.get('confidence_reason', '')),
            ])
        try:
            t = Table(data, hAlign="LEFT", repeatRows=1)
            t.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
            ]))
            story.append(t)
            story.append(Spacer(1, 12))
        except Exception:
            story.append(Paragraph("Unable to render actionable items table in PDF", styles["Normal"]))

    # Resources
    story.append(Paragraph("Resource Inventory", styles["Heading2"]))
    res_data = [["Resource", "Count"]]
    for k, v in resources.items():
        res_data.append([k, str(v)])
    rt = Table(res_data, hAlign="LEFT")
    rt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
    story.append(rt)
    story.append(Spacer(1, 12))

    # Top resources
    if top_resources:
        story.append(Paragraph("Top Resources (by cost)", styles["Heading2"]))
        tr_data = [["ResourceId", "Amount (USD)"]]
        for rid, amt in top_resources:
            tr_data.append([rid, f"${amt:.2f}"])
        trt = Table(tr_data, hAlign="LEFT")
        trt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
        story.append(trt)
        story.append(Spacer(1, 12))

    # Tag aggregates
    if tag_aggregates:
        story.append(Paragraph("Tag aggregates", styles["Heading2"]))
        for tag, mapping in tag_aggregates.items():
            story.append(Paragraph(f"Tag: {tag}", styles["Heading3"]))
            data = [["TagValue", "Amount (USD)"]]
            for tv, amt in mapping.items():
                data.append([tv, f"${amt:.2f}"])
            tt = Table(data, hAlign="LEFT")
            tt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
            story.append(tt)
            story.append(Spacer(1, 6))

    # Add charts
    story.append(Paragraph("Charts", styles["Heading2"]))
    try:
        if cost_plot:
            story.append(RLImage(cost_plot, width=450, height=200))
            story.append(Spacer(1, 6))
        if stacked_area:
            story.append(RLImage(stacked_area, width=450, height=200))
            story.append(Spacer(1, 6))
        if top_bar:
            story.append(RLImage(top_bar, width=450, height=200))
            story.append(Spacer(1, 6))
        if forecast_plot:
            story.append(RLImage(forecast_plot, width=450, height=200))
            story.append(Spacer(1, 6))
    except Exception:
        story.append(Paragraph("Unable to attach images to PDF", styles["Normal"]))

    # Wasted resources
    if wasted:
        story.append(Paragraph("Potential wasted resources", styles["Heading2"]))
        for k, lst in wasted.items():
            story.append(Paragraph(f"{k} ({len(lst)}):", styles["Heading3"]))
            # show up to 50 items
            max_show = 50
            data = [["ResourceId"]]
            for item in lst[:max_show]:
                data.append([str(item)])
            ttt = Table(data, hAlign="LEFT")
            ttt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
            story.append(ttt)
            story.append(Spacer(1, 6))

    # Add images
    try:
        story.append(Paragraph("Graphs", styles["Heading2"]))
        story.append(Spacer(1, 6))
        story.append(RLImage(cost_plot, width=450, height=200))
        story.append(Spacer(1, 6))
        story.append(RLImage(service_pie, width=300, height=300))
    except Exception:
        story.append(Paragraph("Unable to attach images to PDF", styles["Normal"]))

    doc.build(story)


def main():
    args = parse_args()
    global _BOTO_SESSION
    if args.profile:
        _BOTO_SESSION = get_boto_session(args.profile)
    else:
        _BOTO_SESSION = get_boto_session()
    analyze(args)


if __name__ == "__main__":
    main()
