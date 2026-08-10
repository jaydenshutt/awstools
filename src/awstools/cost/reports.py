"""HTML/PDF/CSV report generation for cost analysis."""

from __future__ import annotations

import csv
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib

matplotlib.use("Agg")  # headless-safe
import matplotlib.pyplot as plt
import pandas as pd
from jinja2 import Template

from awstools.branding import (
    ATTRIBUTION_HTML,
    ATTRIBUTION_PLAIN,
    AUTHOR_URL,
    NO_WARRANTY_HTML,
    NO_WARRANTY_PLAIN,
    NO_WARRANTY_SHORT,
)
from awstools.cost.costs import forecast_linear
from awstools.cost.recommendations import compute_priority_confidence

LOG = logging.getLogger("awstools.cost.reports")

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        Image as RLImage,
    )

    HAS_REPORTLAB = True
except Exception:
    HAS_REPORTLAB = False
    letter = None


REPORT_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>AWS Cost Report - {{ account_id }}</title>
  <style>
    :root {
      --bg: #0b0f14;
      --card: #141b24;
      --card2: #1a2330;
      --text: #e8eef6;
      --muted: #8b9bb0;
      --accent: #3b82f6;
      --ok: #22c55e;
      --warn: #f59e0b;
      --danger: #ef4444;
      --border: #243041;
      --chip: #0f1720;
    }
    * { box-sizing: border-box; }
    body {
      font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
      margin: 0; padding: 0; background: var(--bg); color: var(--text); line-height: 1.55;
    }
    .wrap { max-width: 1100px; margin: 0 auto; padding: 28px 20px 48px; }
    .hero {
      background: linear-gradient(145deg, #152033 0%, #0f141c 60%);
      border: 1px solid var(--border); border-radius: 16px; padding: 24px 26px;
      margin-bottom: 20px;
    }
    .hero h1 { margin: 0 0 6px; font-size: 1.65rem; letter-spacing: -0.02em; }
    .hero .sub { color: var(--muted); font-size: 0.95rem; }
    .badge {
      display: inline-block; font-size: 0.72rem; text-transform: uppercase;
      letter-spacing: 0.06em; padding: 3px 8px; border-radius: 999px;
      background: var(--chip); border: 1px solid var(--border); color: var(--muted);
      margin-left: 8px; vertical-align: middle;
    }
    .badge.warn { color: var(--warn); border-color: #5c4a1f; }
    .badge.ok { color: var(--ok); border-color: #1f4d32; }
    .kpis {
      display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
      gap: 12px; margin-top: 18px;
    }
    .kpi {
      background: var(--card); border: 1px solid var(--border);
      border-radius: 12px; padding: 14px 16px;
    }
    .kpi label {
      display: block; color: var(--muted); font-size: 0.72rem;
      text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px;
    }
    .kpi .v { font-size: 1.35rem; font-weight: 700; }
    .kpi .hint { color: var(--muted); font-size: 0.8rem; margin-top: 2px; }
    nav.toc {
      display: flex; flex-wrap: wrap; gap: 8px; margin: 16px 0 22px;
    }
    nav.toc a {
      color: var(--muted); text-decoration: none; font-size: 0.85rem;
      border: 1px solid var(--border); padding: 6px 10px; border-radius: 999px;
      background: var(--card);
    }
    nav.toc a:hover { color: var(--text); border-color: var(--accent); }
    section {
      background: var(--card); border: 1px solid var(--border);
      border-radius: 14px; padding: 20px 22px; margin-bottom: 16px;
    }
    section.engineer { background: var(--card2); }
    h2 {
      margin: 0 0 12px; font-size: 1.05rem; letter-spacing: -0.01em;
      display: flex; align-items: center; gap: 8px;
    }
    h2 .num {
      font-size: 0.75rem; color: var(--muted); border: 1px solid var(--border);
      border-radius: 6px; padding: 1px 6px;
    }
    h3 { margin: 14px 0 8px; font-size: 0.92rem; color: var(--muted); font-weight: 600; }
    p { margin: 0 0 10px; }
    .healthy {
      border-left: 4px solid var(--ok); background: #0f1f18; padding: 12px 14px;
      border-radius: 0 10px 10px 0; margin: 8px 0;
    }
    .alert {
      border-left: 4px solid var(--warn); background: #1f1a10; padding: 12px 14px;
      border-radius: 0 10px 10px 0; margin: 8px 0;
    }
    .actions { display: grid; gap: 10px; }
    .action {
      display: grid; grid-template-columns: 72px 1fr auto; gap: 12px; align-items: start;
      background: #0f1520; border: 1px solid var(--border); border-radius: 10px; padding: 12px 14px;
    }
    .action .rank {
      font-weight: 700; color: var(--accent); font-size: 0.85rem; padding-top: 2px;
    }
    .action .title { font-weight: 600; margin-bottom: 2px; }
    .action .meta { color: var(--muted); font-size: 0.85rem; }
    .action .usd { color: var(--ok); font-weight: 700; white-space: nowrap; }
    .charts {
      display: grid; grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); gap: 14px;
    }
    .chart {
      background: #0f1520; border: 1px solid var(--border); border-radius: 10px; padding: 10px;
    }
    .chart h3 { margin: 0 0 8px; }
    .chart img { max-width: 100%; height: auto; border-radius: 6px; background: #fff; }
    table { width: 100%; border-collapse: collapse; font-size: 0.9rem; }
    th, td { text-align: left; padding: 8px 6px; border-bottom: 1px solid var(--border); }
    th { color: var(--muted); font-weight: 600; font-size: 0.78rem; text-transform: uppercase; }
    ul { margin: 0; padding-left: 1.15rem; }
    li { margin: 5px 0; }
    code, pre {
      font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
      font-size: 0.8rem;
    }
    pre {
      background: #0b111a; padding: 12px; border-radius: 8px; overflow: auto;
      max-height: 280px; border: 1px solid var(--border);
    }
    details { margin-top: 8px; }
    summary { cursor: pointer; color: var(--muted); }
    footer {
      margin-top: 24px; color: var(--muted); font-size: 0.82rem;
      border-top: 1px solid var(--border); padding-top: 14px;
    }
    a { color: var(--accent); }
    @media print {
      body { background: #fff; color: #111; }
      .hero, section { border-color: #ddd; background: #fff; }
      nav.toc { display: none; }
    }
  </style>
</head>
<body>
<div class="wrap">

  <header class="hero">
    <h1>
      AWS cost report
      {% if ce_error or (warnings and warnings|length) %}
        <span class="badge warn">partial coverage</span>
      {% else %}
        <span class="badge ok">ready</span>
      {% endif %}
    </h1>
    <div class="sub">
      Account <strong>{{ account_id }}</strong>
      · Period <strong>{{ start }} - {{ end }}</strong>
      · awstools
      · {{ attribution_plain }}
    </div>
    <div class="kpis" aria-label="executive-kpis">
      <div class="kpi">
        <label>Latest monthly spend</label>
        <div class="v">${{ '%.2f'|format(latest_total) }}</div>
        <div class="hint">Unblended Cost Explorer</div>
      </div>
      <div class="kpi">
        <label>Top cost driver</label>
        <div class="v" style="font-size:1rem">
          {% if top_services %}{{ top_services[0][0][:36] }}{% else %}-{% endif %}
        </div>
        <div class="hint">
          {% if top_services %}${{ '%.2f'|format(top_services[0][1]) }}/mo{% else %}n/a{% endif %}
        </div>
      </div>
      <div class="kpi">
        <label>Priority actions</label>
        <div class="v">{{ rec_objects|length if rec_objects else 0 }}</div>
        <div class="hint">Review before acting</div>
      </div>
      <div class="kpi">
        <label>Anomalies (MoM)</label>
        <div class="v">{{ anomalies|length if anomalies else 0 }}</div>
        <div class="hint">≥ threshold moves</div>
      </div>
      <div class="kpi">
        <label>Regions / errors</label>
        <div class="v">{{ regions_scanned }} / {{ error_count }}</div>
        <div class="hint">Inventory coverage</div>
      </div>
    </div>
  </header>

  <nav class="toc" aria-label="sections">
    <a href="#exec">Executive</a>
    <a href="#actions">Top actions</a>
    <a href="#charts">Charts</a>
    <a href="#anomalies">Anomalies</a>
    <a href="#engineer">Engineer detail</a>
  </nav>

  <section id="exec">
    <h2><span class="num">01</span> Executive summary</h2>
    <p>{{ executive_summary }}</p>
    {% if not rec_objects and not anomalies %}
      <div class="healthy">
        <strong>No urgent action flagged from available data.</strong>
        Re-run after Cost Explorer is fully populated, or expand region coverage.
      </div>
    {% endif %}
    {% if ce_error %}
      <div class="alert"><strong>Cost Explorer:</strong> {{ ce_error }}</div>
    {% endif %}
    {% if warnings %}
      <h3>Coverage notes</h3>
      <ul>{% for w in warnings %}<li>{{ w }}</li>{% endfor %}</ul>
    {% endif %}
    {% if commitments and commitments.hint %}
      <h3>Commitments (directional)</h3>
      <p>{{ commitments.hint }}</p>
      {% if commitments.on_demand_share_pct is defined %}
      <p class="sub" style="color:var(--muted)">
        On-Demand ~{{ commitments.on_demand_share_pct }}% ·
        Reserved ~{{ commitments.reserved_share_pct }}% ·
        Savings Plans ~{{ commitments.savings_plan_share_pct }}%
      </p>
      {% endif %}
    {% endif %}
    {% if ce_forecast and ce_forecast.available %}
      <h3>CE forecast</h3>
      <p>Mean forecast total:
        <strong>${{ '%.2f'|format(ce_forecast.mean_value) }}</strong>
        ({{ ce_forecast.start }} → {{ ce_forecast.end }})
      </p>
    {% endif %}
  </section>

  <section id="actions">
    <h2><span class="num">02</span> Top actions</h2>
    {% if rec_objects %}
    <div class="actions">
      {% for o in rec_objects[:8] %}
      <div class="action">
        <div class="rank">#{{ loop.index }}</div>
        <div>
          <div class="title">{{ o.service }}</div>
          <div class="meta">{{ o.note }}</div>
        </div>
        <div class="usd">${{ '%.2f'|format(o.estimated_monthly_savings) }}/mo</div>
      </div>
      {% endfor %}
    </div>
    <p style="color:var(--muted);font-size:0.85rem;margin-top:12px">
      Dollar figures are conservative heuristics unless backed by CUR / measured sizes.
      Validate before purchasing commitments or deleting resources.
    </p>
    <h3>How estimates work</h3>
    <ul style="color:var(--muted);font-size:0.88rem">
      <li><strong>High confidence</strong> waste (e.g. unattached EBS, idle EIP) uses measured or well-known unit prices.</li>
      <li><strong>Service recommendations</strong> use a conservative % of that service’s Cost Explorer spend (not the whole bill).</li>
      <li><strong>Medium / Low</strong> items (rightsizing, old AMIs, NAT review) are for human review - not auto-action.</li>
      <li>Re-run <code>waste</code> for resource-level findings with <code>estimate_formula</code> on each row.</li>
    </ul>
    {% else %}
      <div class="healthy">No ranked recommendations for this period.</div>
    {% endif %}
  </section>

  <section id="charts">
    <h2><span class="num">03</span> Charts</h2>
    <div class="charts">
      <div class="chart"><h3>Trend</h3><img src="{{ cost_plot }}" alt="Total cost trend"></div>
      <div class="chart"><h3>Service share</h3><img src="{{ service_pie }}" alt="Service share"></div>
      <div class="chart"><h3>Stacked services</h3><img src="{{ stacked_area }}" alt="Stacked area"></div>
      <div class="chart"><h3>Top services</h3><img src="{{ top_bar }}" alt="Top services"></div>
      <div class="chart"><h3>Forecast</h3><img src="{{ forecast_plot }}" alt="Forecast"></div>
    </div>
  </section>

  <section id="anomalies">
    <h2><span class="num">04</span> Anomalies &amp; top services</h2>
    {% if anomalies %}
      <ul>
      {% for a in anomalies[:12] %}
        <li>{{ a.message }}</li>
      {% endfor %}
      </ul>
    {% else %}
      <p style="color:var(--muted)">No MoM anomalies above threshold.</p>
    {% endif %}
    <h3>Latest month services</h3>
    <table>
      <thead><tr><th>Service</th><th>Amount (USD)</th></tr></thead>
      <tbody>
      {% for svc, amt in top_services %}
        <tr><td>{{ svc }}</td><td>${{ '%.2f'|format(amt) }}</td></tr>
      {% endfor %}
      </tbody>
    </table>
  </section>

  <section id="engineer" class="engineer">
    <h2><span class="num">05</span> Engineer detail</h2>

    <h3>Full recommendation text</h3>
    {% for r in recommendations %}
      <div class="action" style="grid-template-columns:1fr"><div class="meta">{{ r }}</div></div>
    {% endfor %}

    <h3>Prescriptive guidance by service</h3>
    {% for svc, info in per_service_advice.items() %}
      <details>
        <summary><strong>{{ svc }}</strong> - est. ${{ '%.2f'|format(info.estimated_monthly_savings) }}/mo</summary>
        <p>{{ info.explanation }}</p>
        <ul>{% for a in info.actions %}<li>{{ a }}</li>{% endfor %}</ul>
      </details>
    {% endfor %}

    <h3>Wasted resources (inventory signals)</h3>
    {% for k, lst in wasted_resources.items() %}
      <details>
        <summary>{{ k }} ({{ lst|length }})</summary>
        <ul>{% for item in lst[:40] %}<li><code>{{ item }}</code></li>{% endfor %}</ul>
      </details>
    {% endfor %}

    <h3>Top resources by cost</h3>
    <ul>
    {% for rid, amt in top_resources %}
      <li><code>{{ rid }}</code>: ${{ '%.2f'|format(amt) }}</li>
    {% else %}
      <li style="color:var(--muted)">No resource-level costs available</li>
    {% endfor %}
    </ul>

    {% if tag_aggregates %}
    <h3>Tag aggregates</h3>
    {% for tag, mapping in tag_aggregates.items() %}
      <h3 style="font-size:0.85rem">Tag: {{ tag }}</h3>
      <ul>{% for tv, v in mapping.items() %}<li>{{ tv }}: ${{ '%.2f'|format(v) }}</li>{% endfor %}</ul>
    {% endfor %}
    {% endif %}

    <h3>Glossary</h3>
    <table>
      <tbody>
      {% for term, desc in glossary.items() %}
        <tr><th style="width:28%">{{ term }}</th><td>{{ desc }}</td></tr>
      {% endfor %}
      </tbody>
    </table>

    <h3>Resource inventory (JSON)</h3>
    <pre>{{ resources_json }}</pre>
  </section>

  <footer>
    Generated by awstools. {{ attribution_html }}.
    Estimates are heuristics; validate before acting.
    Prefer <code>waste</code> / cleanup dry-runs before any <code>--execute</code>.
    <p style="margin-top:10px">{{ no_warranty_html }}</p>
  </footer>
</div>
</body>
</html>
"""


def generate_plots(pivot: pd.DataFrame, out_dir: Path) -> Dict[str, str]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = {
        "cost_plot": str(out_dir / "costs.png"),
        "service_pie": str(out_dir / "top_services.png"),
        "stacked_area": str(out_dir / "stacked_area.png"),
        "top_bar": str(out_dir / "top_services_bar.png"),
        "forecast_plot": str(out_dir / "forecast.png"),
    }

    def placeholder(path: str, msg: str = "No data"):
        fig = plt.figure(figsize=(6, 3))
        plt.text(0.5, 0.5, msg, ha="center", va="center")
        plt.axis("off")
        fig.savefig(path, bbox_inches="tight")
        plt.close(fig)

    if pivot.empty:
        for p in paths.values():
            placeholder(p)
        return paths

    fig, ax = plt.subplots(figsize=(10, 4))
    pivot["total"].plot(ax=ax, marker="o")
    ax.set_title("Total monthly cost")
    ax.set_ylabel("USD")
    fig.tight_layout()
    fig.savefig(paths["cost_plot"])
    plt.close(fig)

    services_only = pivot.drop(columns=["total"]) if "total" in pivot.columns else pivot
    if not services_only.empty:
        fig3, ax3 = plt.subplots(figsize=(10, 4))
        services_only.plot(kind="area", stacked=True, ax=ax3, legend=False)
        ax3.set_title("Monthly cost by service (stacked)")
        ax3.set_ylabel("USD")
        fig3.tight_layout()
        fig3.savefig(paths["stacked_area"])
        plt.close(fig3)
    else:
        placeholder(paths["stacked_area"])

    last = pivot.iloc[-1].drop(labels=["total"]).sort_values(ascending=False)
    # Collapse long tails for pie readability
    if len(last) > 8:
        top = last.head(7)
        other = last.iloc[7:].sum()
        last_pie = pd.concat([top, pd.Series({"Other": other})])
    else:
        last_pie = last

    fig2, ax2 = plt.subplots(figsize=(6, 6))
    if last_pie.sum() > 0:
        ax2.pie(last_pie, labels=last_pie.index, autopct="%.1f%%", textprops={"fontsize": 8})
        ax2.set_title("Cost share by service (latest)")
    else:
        ax2.text(0.5, 0.5, "No cost data", ha="center", va="center")
    fig2.tight_layout()
    fig2.savefig(paths["service_pie"])
    plt.close(fig2)

    topn = last.head(10)
    fig4, ax4 = plt.subplots(figsize=(8, 4))
    topn.plot(kind="bar", ax=ax4)
    ax4.set_title("Top services (latest month)")
    ax4.set_ylabel("USD")
    fig4.tight_layout()
    fig4.savefig(paths["top_bar"])
    plt.close(fig4)

    try:
        series = pivot["total"]
        preds = forecast_linear(series, months=3)
        fig5, ax5 = plt.subplots(figsize=(10, 4))
        series.plot(ax=ax5, marker="o", label="Historical")
        preds.plot(ax=ax5, marker="x", linestyle="--", label="Forecast")
        ax5.set_title("Historical and forecasted total cost")
        ax5.set_ylabel("USD")
        ax5.legend()
        fig5.tight_layout()
        fig5.savefig(paths["forecast_plot"])
        plt.close(fig5)
    except Exception:
        placeholder(paths["forecast_plot"], "No forecast data")

    return paths


def write_html_report(
    html_path: Path,
    context: Dict[str, Any],
) -> None:
    html_path.parent.mkdir(parents=True, exist_ok=True)
    # Make image paths relative to HTML location
    html_dir = html_path.parent
    for key in (
        "cost_plot",
        "service_pie",
        "stacked_area",
        "top_bar",
        "forecast_plot",
    ):
        if context.get(key):
            try:
                context[key] = os.path.relpath(context[key], html_dir)
            except ValueError:
                pass
    context.setdefault("attribution_plain", ATTRIBUTION_PLAIN)
    context.setdefault("attribution_html", ATTRIBUTION_HTML)
    context.setdefault("author_url", AUTHOR_URL)
    context.setdefault("no_warranty_html", NO_WARRANTY_HTML)
    context.setdefault("no_warranty_plain", NO_WARRANTY_PLAIN)
    context.setdefault("no_warranty_short", NO_WARRANTY_SHORT)
    html = Template(REPORT_TEMPLATE).render(**context)
    html_path.write_text(html, encoding="utf-8")
    LOG.info("HTML report written to %s", html_path)


def write_actions_export(
    path: Path,
    rec_objects: List[Dict],
    wasted: Dict[str, List[str]],
    latest_total: float,
) -> List[Dict[str, str]]:
    """Write actionable CSV (+ XLSX) and return rows for PDF."""
    rows: List[Dict[str, str]] = []
    for o in rec_objects or []:
        est = float(o.get("estimated_monthly_savings", 0.0) or 0.0)
        source = o.get("source", "heuristic")
        extra = {"count": o["count"]} if "count" in o else o.get("id")
        pr, conf, reason, score = compute_priority_confidence(
            est, latest_total, source=source, extra=extra
        )
        rows.append(
            {
                "resource": o.get("service") or "account",
                "action": o.get("note") or "recommendation",
                "estimated_monthly_saving": f"{est:.2f}",
                "priority": pr,
                "priority_score": f"{score:.2f}",
                "confidence": conf,
                "confidence_reason": reason,
                "notes": source,
            }
        )

    unit_est = {
        "unattached_ebs": 5.0,
        "unassociated_eips": 3.65,
        "old_available_snapshots": 1.0,
    }
    for k, lst in (wasted or {}).items():
        if k == "available_nat_gateways":
            continue  # review-only
        est_val = unit_est.get(k, 0.0)
        for item in lst:
            pr, conf, reason, score = compute_priority_confidence(
                est_val, latest_total, source="wasted", extra={"count": len(lst)}
            )
            rows.append(
                {
                    "resource": str(item),
                    "action": f"review and/or remediate ({k})",
                    "estimated_monthly_saving": f"{est_val:.2f}" if est_val else "",
                    "priority": pr,
                    "priority_score": f"{score:.2f}",
                    "confidence": conf,
                    "confidence_reason": reason,
                    "notes": f"wasted resource: {k}",
                }
            )

    path = Path(path)
    if path.suffix.lower() != ".csv":
        path = path.with_suffix(".csv")
    path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "resource",
        "action",
        "estimated_monthly_saving",
        "priority",
        "priority_score",
        "confidence",
        "confidence_reason",
        "notes",
    ]
    with path.open("w", newline="", encoding="utf-8") as cf:
        writer = csv.DictWriter(cf, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow({fn: r.get(fn, "") for fn in fieldnames})
    LOG.info("Actionable CSV written to %s", path)

    try:
        xlsx_path = path.with_suffix(".xlsx")
        df = pd.DataFrame(rows).reindex(columns=fieldnames)
        df.to_excel(xlsx_path, index=False)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import numbers

            wb = load_workbook(xlsx_path)
            ws = wb.active
            header = [c.value for c in ws[1]]
            if "estimated_monthly_saving" in header:
                col_idx = header.index("estimated_monthly_saving") + 1
                for row in ws.iter_rows(
                    min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row
                ):
                    for cell in row:
                        try:
                            if cell.value not in (None, ""):
                                cell.value = float(cell.value)
                                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        except Exception:
                            pass
            wb.save(xlsx_path)
        except Exception:
            pass
        LOG.info("Actionable XLSX written to %s", xlsx_path)
    except Exception as e:
        LOG.warning("XLSX export failed: %s", e)

    return rows


def write_pdf_report(
    pdf_path: Path,
    start: str,
    end: str,
    latest_total: float,
    top_services: List[Tuple[str, float]],
    recommendations: List[str],
    plots: Dict[str, str],
    resources: Dict[str, Any],
    top_resources: Optional[List] = None,
    tag_aggregates: Optional[Dict] = None,
    wasted: Optional[Dict] = None,
    rec_objects: Optional[List] = None,
    executive_summary: Optional[str] = None,
    actionable_rows: Optional[List] = None,
    account_id: Optional[str] = None,
) -> None:
    if not HAS_REPORTLAB:
        raise RuntimeError("reportlab is required for PDF export")

    pdf_path = Path(pdf_path)
    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("AWS Account Cost Analysis (awstools)", styles["Title"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(ATTRIBUTION_PLAIN, styles["Normal"]))
    story.append(Paragraph(NO_WARRANTY_SHORT, styles["Normal"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Period: {start} to {end}", styles["Normal"]))
    story.append(
        Paragraph(f"Latest monthly cost: ${latest_total:.2f}", styles["Normal"])
    )
    if account_id:
        story.append(Paragraph(f"Account: {account_id}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if executive_summary:
        story.append(Paragraph("Executive summary", styles["Heading2"]))
        story.append(Paragraph(executive_summary, styles["Normal"]))
        story.append(Spacer(1, 12))

    story.append(Paragraph("Top Services", styles["Heading2"]))
    data = [["Service", "Amount (USD)"]]
    for svc, amt in top_services:
        data.append([str(svc)[:60], f"${amt:.2f}"])
    t = Table(data, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Recommendations", styles["Heading2"]))
    for r in recommendations:
        story.append(Paragraph(f"• {r}", styles["Normal"]))
        story.append(Spacer(1, 4))

    if rec_objects:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Estimated savings", styles["Heading2"]))
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, leading=11)
        data = [
            [
                Paragraph("<b>Service</b>", small),
                Paragraph("<b>Note</b>", small),
                Paragraph("<b>Est. $/mo</b>", small),
            ]
        ]
        for o in rec_objects:
            data.append(
                [
                    Paragraph(str(o.get("service") or ""), small),
                    Paragraph(str(o.get("note") or "")[:200], small),
                    Paragraph(f"${o.get('estimated_monthly_savings', 0):.2f}", small),
                ]
            )
        t = Table(data, colWidths=[110, 320, 80], hAlign="LEFT")
        t.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(t)

    # Resource counts only (not full lists) for PDF readability
    story.append(Spacer(1, 12))
    story.append(Paragraph("Resource inventory (counts)", styles["Heading2"]))
    res_data = [["Resource", "Count"]]
    for k, v in resources.items():
        if k.startswith("_") or k.endswith("_list"):
            continue
        res_data.append([k, str(v)])
    rt = Table(res_data, hAlign="LEFT")
    rt.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey)]))
    story.append(rt)

    if wasted:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Potential wasted resources", styles["Heading2"]))
        for k, lst in wasted.items():
            story.append(Paragraph(f"{k}: {len(lst)} item(s)", styles["Normal"]))

    story.append(Spacer(1, 12))
    story.append(Paragraph("Charts", styles["Heading2"]))
    for key in ("cost_plot", "stacked_area", "top_bar", "forecast_plot", "service_pie"):
        p = plots.get(key)
        if p and Path(p).exists():
            try:
                w, h = (300, 300) if "pie" in key or "service_pie" in key else (450, 200)
                story.append(RLImage(p, width=w, height=h))
                story.append(Spacer(1, 6))
            except Exception:
                pass

    if actionable_rows:
        story.append(Paragraph("Actionable items", styles["Heading2"]))
        try:
            sorted_rows = sorted(
                actionable_rows,
                key=lambda r: float(r.get("priority_score") or 0),
                reverse=True,
            )
        except Exception:
            sorted_rows = actionable_rows
        data = [["Resource", "Action", "Est $", "Priority", "Confidence"]]
        for r in sorted_rows[:100]:
            data.append(
                [
                    str(r.get("resource", ""))[:40],
                    str(r.get("action", ""))[:50],
                    str(r.get("estimated_monthly_saving", "")),
                    str(r.get("priority", "")),
                    str(r.get("confidence", "")),
                ]
            )
        t = Table(data, hAlign="LEFT", repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.append(t)

    doc.build(story)
    LOG.info("PDF written to %s", pdf_path)
